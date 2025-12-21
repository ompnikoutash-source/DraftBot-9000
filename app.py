import os
import re
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import altair as alt
import openpyxl

# Optional: draggable draft order in sidebar
try:
    from streamlit_sortables import sort_items  # type: ignore
    HAS_SORTABLES = True
except Exception:
    HAS_SORTABLES = False


st.set_page_config(page_title="DraftBot 9000", layout="wide")

POSITIONS = ["QB", "RB", "WR", "TE", "K", "DEF"]
DEFAULT_DRAFT_ORDER = [
    "JAMES", "ERIC", "JACK", "TITO", "KYLE", "BEN",
    "DIEGO", "NIKO", "COLBY", "ALBERTO", "ROBERT", "NOAH"
]
MAX_PICKS = 168  # 12 teams * 14 rounds
MC_SIMS_DEFAULT = 15
MC_CANDIDATES_DEFAULT = 6



# -----------------------------
# Small compatibility helpers
# -----------------------------
def st_df(df: pd.DataFrame, **kwargs):
    """Streamlit dataframe with forward-compatible width handling."""
    try:
        return st.dataframe(df, width="stretch", **kwargs)  # newer Streamlit
    except TypeError:
        return st.dataframe(df, use_container_width=True, **kwargs)  # older Streamlit


def st_altair(chart, **kwargs):
    """Streamlit altair chart with forward-compatible width handling."""
    try:
        return st.altair_chart(chart, width="stretch", **kwargs)  # newer Streamlit
    except TypeError:
        return st.altair_chart(chart, use_container_width=True, **kwargs)  # older Streamlit
def mc_on_clock_key(
    draft_log: pd.DataFrame,
    current_pick: int,
    my_owner: str,
    draft_order: List[str],
    sims: int,
    candidates: int,
) -> str:
    ids = draft_log["player_id"].astype(str).tolist() if not draft_log.empty else []
    return f"{current_pick}|{my_owner}|{sims}|{candidates}|{','.join(draft_order)}|" + ",".join(ids)


    



# -----------------------------
# League settings and draft math
# -----------------------------
@dataclass
class LeagueSettings:
    num_teams: int = 12
    starters: Dict[str, int] = None
    run_window: int = 8
    history_year: int = 2025

    def __post_init__(self):
        if self.starters is None:
            # QB, WR, WR, RB, RB, W/R, TE, K, DEF
            self.starters = {
                "QB": 1,
                "RB": 2,
                "WR": 2,
                "TE": 1,
                "FLEX_RBWR": 1,
                "K": 1,
                "DEF": 1,
            }


def snake_team_index(pick_number: int, num_teams: int) -> int:
    round_number = (pick_number - 1) // num_teams + 1
    within = (pick_number - 1) % num_teams
    if round_number % 2 == 1:
        return within
    return num_teams - 1 - within


def next_pick_number(draft_log: pd.DataFrame) -> int:
    if draft_log.empty:
        return 1
    return int(draft_log["pick"].max()) + 1


def pick_to_round(pick: int, num_teams: int) -> int:
    return (pick - 1) // num_teams + 1


def _clean_raw_name(raw: str) -> str:
    return str(raw).replace("View News", "").strip()

def next_pick_for_owner(start_pick: int, draft_order: List[str], owner: str, num_teams: int, max_picks: int) -> Optional[int]:
    p = start_pick
    while p <= max_picks:
        idx = snake_team_index(p, num_teams)
        if draft_order[idx] == owner:
            return p
        p += 1
    return None



# -----------------------------
# Workbook parsing (updated)
# -----------------------------
def _find_section_header_rows(ws) -> Dict[str, int]:
    """
    Finds the header row index in column A for each position section.
    In your workbook, sections begin with a row where A == "QB", "RB", "WR", "TE", "K", "DEF".
    """
    header_rows: Dict[str, int] = {}
    for pos in POSITIONS:
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip().upper() == pos:
                header_rows[pos] = r
                break
    return header_rows


def _header_map(ws, header_row: int) -> Dict[int, str]:
    """Maps column index -> header text for a given header row."""
    m: Dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            m[c] = s
    return m


def _choose_points_col(header_by_col: Dict[int, str]) -> Optional[int]:
    """
    Chooses the column that contains Fantasy Points inside a section.
    We prefer exact header "Fantasy Points" but also accept headers containing it.
    """
    for c, h in header_by_col.items():
        if str(h).strip().lower() == "fantasy points":
            return c
    for c, h in header_by_col.items():
        if "fantasy points" in str(h).strip().lower():
            return c
    return None


def parse_player_projections_from_workbook(xlsx_path: str) -> pd.DataFrame:
    """
    Supports two workbook layouts:

    Layout A (old):
      Sheet named "Player Projections" with section headers in col A: QB/RB/WR/TE/K/DEF,
      and each section has its own header row including "Fantasy Points".

    Layout B (new, your current):
      Separate sheets named QB/RB/WR/TE/K/DEF.
      In each sheet the first column header equals the sheet name (ex: "QB"),
      and projection is in "Fantasy Points" (fallback: "Copied Fantasy Points").

    Output columns:
      player, pos, team, proj_points, player_id, available
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = set(wb.sheetnames)

    # -------------------------
    # Layout B: separate tabs
    # -------------------------
    pos_tabs = [p for p in POSITIONS if p in sheet_names and p != "FLEX"]
    if len(pos_tabs) >= 4 and "Player Projections" not in sheet_names:
        rows = []
        for pos in pos_tabs:
            ws = wb[pos]

            # Read header row 1 into a map: col_idx -> header text
            header_by_col = _header_map(ws, 1)

            # Find the player name column: header equals the position name
            name_col = None
            for c, h in header_by_col.items():
                if str(h).strip().upper() == pos:
                    name_col = c
                    break
            if name_col is None:
                name_col = 1

            # Find points column
            pts_col = None
            for c, h in header_by_col.items():
                if str(h).strip().lower() == "fantasy points":
                    pts_col = c
                    break
            if pts_col is None:
                for c, h in header_by_col.items():
                    if str(h).strip().lower() == "copied fantasy points":
                        pts_col = c
                        break
            if pts_col is None:
                raise ValueError(f"Sheet '{pos}' missing 'Fantasy Points' (and no 'Copied Fantasy Points' fallback).")

            for r in range(2, ws.max_row + 1):
                raw_name = ws.cell(r, name_col).value
                if raw_name is None:
                    continue

                name = _clean_raw_name(raw_name)
                if not name:
                    continue

                pts = ws.cell(r, pts_col).value
                try:
                    proj_points = float(pts)
                except Exception:
                    continue
                if np.isnan(proj_points):
                    continue

                team = ""  # your new tabs do not include team as a separate column

                player_id = (
                    re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
                    + "_"
                    + pos
                )

                rows.append(
                    {
                        "player": name,
                        "pos": pos,
                        "team": team,
                        "proj_points": proj_points,
                        "player_id": player_id,
                    }
                )

        pool = pd.DataFrame(rows).drop_duplicates(subset=["player_id"]).reset_index(drop=True)
        pool["available"] = True
        return pool

    # -------------------------
    # Layout A: sectioned sheet
    # -------------------------
    if "Player Projections" in sheet_names:
        ws = wb["Player Projections"]

        header_rows = _find_section_header_rows(ws)
        missing = [p for p in POSITIONS if p not in header_rows]
        if missing:
            raise ValueError(f"Could not find section header rows for: {missing}")

        sections = sorted([(pos, r) for pos, r in header_rows.items()], key=lambda x: x[1])
        ranges: Dict[str, Tuple[int, int]] = {}
        for i, (pos, r0) in enumerate(sections):
            r1 = (sections[i + 1][1] - 1) if i + 1 < len(sections) else ws.max_row
            ranges[pos] = (r0, r1)

        rows = []
        for pos, (r0, r1) in ranges.items():
            header_by_col = _header_map(ws, r0)
            pts_col = _choose_points_col(header_by_col)
            if pts_col is None and header_by_col:
                pts_col = max(header_by_col.keys())

            for r in range(r0 + 1, r1 + 1):
                raw = ws.cell(r, 1).value
                if raw is None:
                    continue

                s = _clean_raw_name(raw)
                if not s:
                    continue

                team = ""
                left = s
                if " - " in s:
                    left, team = s.split(" - ", 1)
                    team = team.strip()

                m = re.search(r"(QB|RB|WR|TE|K|DEF)$", left.strip())
                name = left.strip()
                if m:
                    name = left[:m.start()].strip()

                pts = ws.cell(r, pts_col).value if pts_col is not None else None
                try:
                    proj_points = float(pts)
                except Exception:
                    continue
                if np.isnan(proj_points):
                    continue

                player_id = (
                    re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
                    + "_"
                    + pos
                    + "_"
                    + re.sub(r"[^a-z0-9]+", "_", str(team).lower()).strip("_")
                )

                rows.append(
                    {
                        "player": name,
                        "pos": pos,
                        "team": team,
                        "proj_points": proj_points,
                        "player_id": player_id,
                    }
                )

        pool = pd.DataFrame(rows).drop_duplicates(subset=["player_id"]).reset_index(drop=True)
        pool["available"] = True
        return pool

    raise ValueError("No recognized projection layout found. Expected 'Player Projections' or position sheets QB/RB/WR/TE/K/DEF.")


DRAFT_RESULTS_SHEETS_WEIGHTS = [
    ("Draft Results 1", 0.40),
    ("Draft Results 2", 0.30),
    ("Draft Results 3", 0.15),
    ("Draft Results 4", 0.10),
    ("Draft Results 5", 0.05),
]

def parse_draft_results_tabs(xlsx_path: str) -> Dict[str, pd.DataFrame]:
    """
    Each sheet is expected to be in this shape:
    Round | TEAM1 | TEAM2 | ... | TEAM12
    where the cell values are positions (QB/RB/WR/TE/K/DEF).
    """
    blocks: Dict[str, pd.DataFrame] = {}

    for sheet_name, _w in DRAFT_RESULTS_SHEETS_WEIGHTS:
        try:
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
        except Exception:
            continue

        df.columns = [str(c).strip() for c in df.columns]

        if "Round" not in df.columns:
            # If the first column is unnamed or not called Round
            df = df.rename(columns={df.columns[0]: "Round"})

        df = df.dropna(subset=["Round"]).copy()
        df["Round"] = pd.to_numeric(df["Round"], errors="coerce")
        df = df.dropna(subset=["Round"]).copy()
        df["Round"] = df["Round"].astype(int)

        # Drop unnamed filler columns
        df = df.loc[:, [c for c in df.columns if not str(c).startswith("Unnamed")]]

        blocks[sheet_name] = df.reset_index(drop=True)

    return blocks



# -----------------------------
# VORP with flex handling
# -----------------------------
def compute_vorp_with_flex(pool: pd.DataFrame, settings: LeagueSettings) -> Tuple[pd.DataFrame, Dict[str, float]]:
    df = pool.copy()

    def nth_points(pos: str, n: int) -> float:
        d = df[df["pos"] == pos].sort_values("proj_points", ascending=False).reset_index(drop=True)
        if len(d) == 0:
            return 0.0
        idx = min(n - 1, len(d) - 1)
        return float(d.loc[idx, "proj_points"])

    num_teams = settings.num_teams
    s = settings.starters

    repl: Dict[str, float] = {}
    repl["QB"] = nth_points("QB", num_teams * s["QB"])
    repl["TE"] = nth_points("TE", num_teams * s["TE"])
    repl["K"] = nth_points("K", num_teams * s["K"])
    repl["DEF"] = nth_points("DEF", num_teams * s["DEF"])

    rb_n = num_teams * s["RB"]
    wr_n = num_teams * s["WR"]
    repl["RB"] = nth_points("RB", rb_n)
    repl["WR"] = nth_points("WR", wr_n)

    rb_sorted = df[df["pos"] == "RB"].sort_values("proj_points", ascending=False).reset_index(drop=True)
    wr_sorted = df[df["pos"] == "WR"].sort_values("proj_points", ascending=False).reset_index(drop=True)

    rb_rem = rb_sorted.iloc[rb_n:] if len(rb_sorted) > rb_n else rb_sorted.iloc[0:0]
    wr_rem = wr_sorted.iloc[wr_n:] if len(wr_sorted) > wr_n else wr_sorted.iloc[0:0]

    flex_pool = pd.concat([rb_rem, wr_rem], ignore_index=True)
    flex_n = num_teams * s["FLEX_RBWR"]
    if len(flex_pool) == 0:
        repl["FLEX_RBWR"] = 0.0
    else:
        flex_pool = flex_pool.sort_values("proj_points", ascending=False).reset_index(drop=True)
        idx = min(flex_n - 1, len(flex_pool) - 1)
        repl["FLEX_RBWR"] = float(flex_pool.loc[idx, "proj_points"])

    def baseline(row) -> float:
        if row["pos"] in ["RB", "WR"]:
            return max(repl[row["pos"]], repl["FLEX_RBWR"])
        return repl.get(row["pos"], 0.0)

    df["replacement_points"] = df.apply(baseline, axis=1)
    df["vorp"] = df["proj_points"] - df["replacement_points"]
    return df, repl

def roster_counts_from_log(draft_log: pd.DataFrame, owner: str) -> Dict[str, int]:
    if draft_log.empty:
        return {}
    mine = draft_log[draft_log["owner"] == owner]
    return mine["pos"].value_counts().to_dict()


def need_bonus_for_owner(owner: str, draft_log: pd.DataFrame, settings: LeagueSettings) -> Dict[str, float]:
    counts = roster_counts_from_log(draft_log, owner)
    targets = {
        "QB": settings.starters["QB"],
        "RB": settings.starters["RB"],
        "WR": settings.starters["WR"],
        "TE": settings.starters["TE"],
        "K": settings.starters["K"],
        "DEF": settings.starters["DEF"],
    }

    bonus = {p: 0.0 for p in POSITIONS}
    for p in POSITIONS:
        deficit = max(0, targets.get(p, 0) - counts.get(p, 0))
        bonus[p] = float(deficit)

    bonus["RB"] += 0.25
    bonus["WR"] += 0.20
    return bonus


def softmax(x: np.ndarray, temp: float = 1.0) -> np.ndarray:
    x = x / max(1e-6, temp)
    x = x - np.max(x)
    ex = np.exp(x)
    return ex / np.sum(ex)


def sample_opponent_pick(avail: pd.DataFrame, owner: str, draft_log: pd.DataFrame, settings: LeagueSettings, temp: float = 0.9) -> str:
    if avail.empty:
        return ""

    bonus = need_bonus_for_owner(owner, draft_log, settings)
    work = avail.copy()
    work["score"] = work["vorp"] + work["pos"].map(bonus)

    work = work.sort_values("score", ascending=False).head(25).reset_index(drop=True)
    probs = softmax(work["score"].to_numpy(dtype=float), temp=temp)
    idx = np.random.choice(len(work), p=probs)
    return str(work.loc[idx, "player_id"])
def simulate_finish(
    pool: pd.DataFrame,
    draft_order: List[str],
    draft_log: pd.DataFrame,
    settings: LeagueSettings,
    start_pick: int,
    forced_my_first_pick_id: Optional[str],
    my_owner: str,
) -> Tuple[float, float]:
    # Build lookup dict keyed by player_id
    by_id = (
        pool.set_index("player_id")[["player", "pos", "team", "proj_points", "vorp"]]
        .to_dict(orient="index")
    )

    taken = set(draft_log["player_id"].tolist()) if not draft_log.empty else set()
    sim_log = draft_log.copy()
    forced_used = False

    def available_df() -> pd.DataFrame:
        if not taken:
            return pool.copy()
        return pool[~pool["player_id"].isin(taken)].copy()

    for p in range(start_pick, MAX_PICKS + 1):
        owner = draft_order[snake_team_index(p, settings.num_teams)]
        av = available_df()
        if av.empty:
            break

        # Force "my pick" to be the candidate under evaluation (first time only)
        if (owner == my_owner) and (forced_my_first_pick_id is not None) and (not forced_used):
            pick_id = forced_my_first_pick_id if (forced_my_first_pick_id in by_id and forced_my_first_pick_id not in taken) else ""
            forced_used = True
            if not pick_id:
                pick_id = str(av.sort_values("vorp", ascending=False).iloc[0]["player_id"])
        else:
            pick_id = sample_opponent_pick(av, owner, sim_log, settings, temp=0.95)
            if not pick_id:
                pick_id = str(av.sort_values("vorp", ascending=False).iloc[0]["player_id"])

        taken.add(pick_id)

        r = by_id[pick_id]
        rnd = pick_to_round(p, settings.num_teams)
        sim_log = pd.concat(
            [sim_log, pd.DataFrame([{
                "pick": p,
                "round": rnd,
                "owner": owner,
                "player_id": pick_id,
                "player": r["player"],
                "pos": r["pos"],
                "team": r["team"],
                "proj_points": float(r["proj_points"]),
                "vorp": float(r["vorp"]),
                "ts": time.time(),
            }])],
            ignore_index=True
        )

    # Score "my" starters after the sim completes
    my = sim_log[sim_log["owner"] == my_owner].copy()

    def take_pos(df_team: pd.DataFrame, pos: str, n: int) -> pd.DataFrame:
        return df_team[df_team["pos"] == pos].sort_values("proj_points", ascending=False).head(n)

    qb = take_pos(my, "QB", 1)
    rb = take_pos(my, "RB", 2)
    wr = take_pos(my, "WR", 2)
    te = take_pos(my, "TE", 1)
    k = take_pos(my, "K", 1)
    d = take_pos(my, "DEF", 1)

    used = set(pd.concat([qb, rb, wr, te, k, d], ignore_index=True)["player_id"].tolist())
    flex_pool = my[my["pos"].isin(["RB", "WR"])].copy()
    flex_pool = flex_pool[~flex_pool["player_id"].isin(used)]
    flex = flex_pool.sort_values("proj_points", ascending=False).head(1)

    starters = pd.concat([qb, rb, wr, te, flex, k, d], ignore_index=True)
    return float(starters["proj_points"].sum()), float(starters["vorp"].sum())

def mc_best_pick_on_clock(
    pool_vorp: pd.DataFrame,
    draft_order: List[str],
    draft_log: pd.DataFrame,
    settings: LeagueSettings,
    current_pick: int,
    my_owner: str,
    sims: int = MC_SIMS_DEFAULT,
    candidates: int = MC_CANDIDATES_DEFAULT,
) -> pd.DataFrame:
    avail_now = pool_vorp[pool_vorp["available"]].copy()
    if avail_now.empty:
        return pd.DataFrame()

    bonus = need_bonus_for_owner(my_owner, draft_log, settings)
    avail_now["pre_score"] = avail_now["vorp"] + avail_now["pos"].map(bonus)
    cand = avail_now.sort_values("pre_score", ascending=False).head(candidates).copy()

    results = []
    for _, row in cand.iterrows():
        pid = str(row["player_id"])
        scores = []

        for _ in range(sims):
            starters_proj, starters_vorp = simulate_finish(
                pool=pool_vorp,
                draft_order=draft_order,
                draft_log=draft_log,
                settings=settings,
                start_pick=current_pick,
                forced_my_first_pick_id=pid,
                my_owner=my_owner,
            )
            score = starters_vorp * 0.65 + starters_proj * 0.35
            scores.append(score)

        results.append({
            "player_id": pid,
            "player": row["player"],
            "pos": row["pos"],
            "team": row["team"],
            "vorp": float(row["vorp"]),
            "proj_points": float(row["proj_points"]),
            "ev_score": float(np.mean(scores)),
            "ev_std": float(np.std(scores)),
            "sims": sims,
        })

    out = pd.DataFrame(results).sort_values("ev_score", ascending=False).reset_index(drop=True)
    out.insert(0, "rank", out.index + 1)
    return out



# -----------------------------
# State and actions
# -----------------------------
def init_state():
    if "xlsx_path" not in st.session_state:
        st.session_state.xlsx_path = "DraftBot 9000.xlsx"

    if "settings" not in st.session_state:
        st.session_state.settings = LeagueSettings()

    if "draft_order" not in st.session_state:
        st.session_state.draft_order = DEFAULT_DRAFT_ORDER.copy()

    if "my_owner" not in st.session_state:
        st.session_state.my_owner = "NIKO"

    if "auto_advance" not in st.session_state:
        st.session_state.auto_advance = True

    if "draft_log" not in st.session_state:
        st.session_state.draft_log = pd.DataFrame(
            columns=["pick", "round", "owner", "player_id", "player", "pos", "team", "proj_points", "vorp", "ts"]
        )

    if "availability" not in st.session_state:
        st.session_state.availability = {}  # player_id -> bool

    if "current_owner" not in st.session_state:
        st.session_state.current_owner = st.session_state.draft_order[0]


@st.cache_data(show_spinner=False)
def load_workbook_data(xlsx_path: str, file_mtime: float):
    pool = parse_player_projections_from_workbook(xlsx_path)
    history_blocks = parse_draft_results_tabs(xlsx_path)
    return pool, history_blocks


def ensure_availability(pool: pd.DataFrame):
    for pid in pool["player_id"].tolist():
        if pid not in st.session_state.availability:
            st.session_state.availability[pid] = True


def undo_last_pick():
    if st.session_state.draft_log.empty:
        return
    last = st.session_state.draft_log.iloc[-1]
    st.session_state.availability[last["player_id"]] = True
    st.session_state.draft_log = st.session_state.draft_log.iloc[:-1].reset_index(drop=True)
    st.session_state.current_owner = str(last["owner"]).strip()


def reset_draft(pool: pd.DataFrame, new_order: List[str]):
    st.session_state.draft_order = new_order
    st.session_state.current_owner = new_order[0]
    st.session_state.draft_log = st.session_state.draft_log.iloc[0:0].copy()
    st.session_state.availability = {pid: True for pid in pool["player_id"].tolist()}


def draft_player(row: pd.Series, owner: str):
    pick = next_pick_number(st.session_state.draft_log)
    if pick > MAX_PICKS:
        return
    if not st.session_state.availability.get(row["player_id"], True):
        return

    rnd = pick_to_round(pick, st.session_state.settings.num_teams)

    st.session_state.draft_log = pd.concat(
        [
            st.session_state.draft_log,
            pd.DataFrame(
                [
                    {
                        "pick": pick,
                        "round": rnd,
                        "owner": owner,
                        "player_id": row["player_id"],
                        "player": row["player"],
                        "pos": row["pos"],
                        "team": row["team"],
                        "proj_points": float(row["proj_points"]),
                        "vorp": float(row["vorp"]),
                        "ts": time.time(),
                    }
                ]
            ),
        ],
        ignore_index=True,
    )

    st.session_state.availability[row["player_id"]] = False

    if st.session_state.auto_advance:
        next_p = pick + 1
        if next_p <= MAX_PICKS:
            idx = snake_team_index(next_p, st.session_state.settings.num_teams)
            st.session_state.current_owner = st.session_state.draft_order[idx]


def roster_counts(owner: str) -> Dict[str, int]:
    dl = st.session_state.draft_log
    mine = dl[dl["owner"] == owner]
    return mine["pos"].value_counts().to_dict()


def run_risk(
    settings: LeagueSettings,
    weighted_histories: List[Tuple[pd.DataFrame, float]],
    draft_log: pd.DataFrame,
    available_df: pd.DataFrame
) -> Dict[str, float]:
    window = settings.run_window
    recent = draft_log.tail(window)
    recent_counts = recent["pos"].value_counts().to_dict()
    denom = max(1, min(window, len(recent)))
    recent_score = {p: recent_counts.get(p, 0) / denom for p in POSITIONS}

    # Weighted historical score for the current round
    current_round = pick_to_round(next_pick_number(draft_log), settings.num_teams)
    hist_score = {p: 0.0 for p in POSITIONS}

    if weighted_histories:
        per_round_scores = []
        weights = []

        for hist_df, w in weighted_histories:
            row = hist_df[hist_df["Round"] == current_round]
            if len(row) != 1:
                continue

            vals = row.iloc[0].drop(labels=["Round"]).astype(str)
            counts = vals.value_counts().to_dict()
            per_round_scores.append({p: counts.get(p, 0) / settings.num_teams for p in POSITIONS})
            weights.append(w)

        if per_round_scores and sum(weights) > 0:
            wsum = sum(weights)
            for p in POSITIONS:
                hist_score[p] = float(
                    sum(s.get(p, 0.0) * w for s, w in zip(per_round_scores, weights)) / wsum
                )

    # Scarcity score
    scarcity = {}
    for p in POSITIONS:
        top = available_df[(available_df["pos"] == p) & (available_df["available"])].nlargest(10, "vorp")
        scarcity[p] = 1.0 - min(len(top) / 10.0, 1.0)

    severity = {}
    for p in POSITIONS:
        severity[p] = float(np.clip(0.45 * recent_score[p] + 0.35 * hist_score[p] + 0.20 * scarcity[p], 0.0, 1.0))
    return severity




def best_available_by_pos(df: pd.DataFrame) -> pd.DataFrame:
    av = df[df["available"]].copy()
    rows = []
    for pos in POSITIONS:
        best = av[av["pos"] == pos].nlargest(1, "vorp")
        if len(best) == 1:
            r = best.iloc[0]
            rows.append(
                {"pos": pos, "player": r["player"], "team": r["team"], "proj_points": r["proj_points"], "vorp": r["vorp"]}
            )
        else:
            rows.append({"pos": pos, "player": None, "team": None, "proj_points": None, "vorp": None})
    return pd.DataFrame(rows)


def need_bonus(my_owner: str, settings: LeagueSettings) -> Dict[str, float]:
    counts = roster_counts(my_owner)
    targets = {
        "QB": settings.starters["QB"],
        "RB": settings.starters["RB"],
        "WR": settings.starters["WR"],
        "TE": settings.starters["TE"],
        "K": settings.starters["K"],
        "DEF": settings.starters["DEF"],
    }

    bonus = {p: 0.0 for p in POSITIONS}
    for p in POSITIONS:
        deficit = max(0, targets.get(p, 0) - counts.get(p, 0))
        bonus[p] = 1.0 * deficit

    bonus["RB"] += 0.25
    bonus["WR"] += 0.20
    return bonus


def compute_team_analysis(draft_log: pd.DataFrame, owners: List[str]) -> pd.DataFrame:
    """
    Draft Analysis: ranks teams using starter strength and total strength (proj and VORP).
    Lineup rules:
      QB1, RB2, WR2, TE1, FLEX(RB/WR)1, K1, DEF1
    """
    if draft_log.empty:
        return pd.DataFrame(columns=[
            "rank", "owner",
            "starters_proj", "starters_vorp",
            "total_proj", "total_vorp",
            "players_drafted"
        ])

    def best_lineup(df_team: pd.DataFrame) -> Tuple[float, float]:
        if df_team.empty:
            return 0.0, 0.0

        def take(pos: str, n: int) -> pd.DataFrame:
            return df_team[df_team["pos"] == pos].sort_values("proj_points", ascending=False).head(n)

        qb = take("QB", 1)
        rb = take("RB", 2)
        wr = take("WR", 2)
        te = take("TE", 1)
        k = take("K", 1)
        d = take("DEF", 1)

        used_ids = set(pd.concat([qb, rb, wr, te, k, d], ignore_index=True)["player_id"].tolist())

        flex_pool = df_team[df_team["pos"].isin(["RB", "WR"])].copy()
        flex_pool = flex_pool[~flex_pool["player_id"].isin(used_ids)]
        flex = flex_pool.sort_values("proj_points", ascending=False).head(1)

        starters = pd.concat([qb, rb, wr, te, flex, k, d], ignore_index=True)
        return float(starters["proj_points"].sum()), float(starters["vorp"].sum())

    rows = []
    for owner in owners:
        dt = draft_log[draft_log["owner"] == owner].copy()
        total_proj = float(dt["proj_points"].sum()) if not dt.empty else 0.0
        total_vorp = float(dt["vorp"].sum()) if not dt.empty else 0.0
        starters_proj, starters_vorp = best_lineup(dt)
        rows.append({
            "owner": owner,
            "starters_proj": starters_proj,
            "starters_vorp": starters_vorp,
            "total_proj": total_proj,
            "total_vorp": total_vorp,
            "players_drafted": int(len(dt)),
        })

    out = pd.DataFrame(rows)
    out["score"] = (out["starters_vorp"] * 0.65) + (out["starters_proj"] * 0.35) + (out["total_vorp"] * 0.05)
    out = out.sort_values("score", ascending=False).reset_index(drop=True)
    out.insert(0, "rank", out.index + 1)
    out = out.drop(columns=["score"])
    return out


# -----------------------------
# App start
# -----------------------------
init_state()
settings: LeagueSettings = st.session_state.settings

# Tight top padding and bubble text sizing
st.markdown(
    """
    <style>
      div[data-testid="stAppViewContainer"] > .main .block-container {
        padding-top: 0.5rem;
        padding-bottom: 1.0rem;
      }
      h1 { margin-top: 0.1rem; padding-top: 0rem; }
      div.stButton > button {
        border-radius: 999px;
        padding: 0.35rem 0.60rem;
        text-align: left;
        width: 100%;
        font-size: 12px;
        line-height: 1.15;
        white-space: normal;
        overflow-wrap: anywhere;
      }
      .slotnum {
        font-weight: 700;
        padding: 0.35rem 0.1rem;
        margin: 0.25rem 0;
        font-size: 13px;
        line-height: 1.1;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

# Sidebar inputs
st.sidebar.header("Inputs")
st.session_state.xlsx_path = st.sidebar.text_input("Workbook filename", st.session_state.xlsx_path)

settings.run_window = st.sidebar.slider("Run detection window (picks)", 4, 16, settings.run_window)
st.session_state.my_owner = st.sidebar.selectbox(
    "My owner",
    DEFAULT_DRAFT_ORDER,
    index=DEFAULT_DRAFT_ORDER.index(st.session_state.my_owner) if st.session_state.my_owner in DEFAULT_DRAFT_ORDER else 0,
)
st.sidebar.subheader("Monte Carlo")
st.session_state.mc_sims = st.sidebar.slider("MC sims", 5, 40, MC_SIMS_DEFAULT)
st.session_state.mc_candidates = st.sidebar.slider("MC candidates", 3, 15, MC_CANDIDATES_DEFAULT)

st.session_state.auto_advance = st.sidebar.checkbox("Auto advance owner (snake)", value=st.session_state.auto_advance)

# Load workbook with cache-busting on file modified time
try:
    mtime = os.path.getmtime(st.session_state.xlsx_path)
except Exception:
    mtime = 0.0

pool, history_blocks = load_workbook_data(st.session_state.xlsx_path, mtime)
ensure_availability(pool)

# Quick sanity warning if any position is missing
counts = pool["pos"].value_counts().to_dict()
missing_pos = [p for p in POSITIONS if counts.get(p, 0) == 0]
if missing_pos:
    st.sidebar.warning(f"No players loaded for: {', '.join(missing_pos)}")

# History year selection for run tendencies
year_options = sorted(list(history_blocks.keys()))
if year_options:
    # Default to the most recent year if nothing is selected yet
    if "history_years" not in st.session_state:
        st.session_state.history_years = [max(year_options)]

    st.session_state.history_years = st.sidebar.multiselect(
        "History years (run tendencies)",
        options=year_options,
        default=st.session_state.history_years,
        help="Select one or more years. The model averages position tendencies across selected years.",
    )

# Draft order in sidebar, draggable blocks with numbers INSIDE the bubbles
with st.sidebar:
    st.subheader("Draft order")
    st.caption("Drag owners to reorder, then click Submit (reset draft).")

    if HAS_SORTABLES:
        custom_style = """
        .sortable-component { padding: 0px; margin: 0px; }
        .sortable-container { padding: 0px; margin: 0px; }
        .sortable-item, .sortable-item:hover {
            border-radius: 999px;
            padding: 0.35rem 0.6rem;
            margin: 0.25rem 0;
            font-size: 13px;
            line-height: 1.1;
        }
        """

        numbered_items = [f"{i+1}. {name}" for i, name in enumerate(st.session_state.draft_order)]
        sorted_items = sort_items(
            numbered_items,
            direction="vertical",
            key="draft_order_sort",
            custom_style=custom_style,
        )

        new_order = [
            re.sub(r"^\s*\d+\.\s*", "", str(x)).strip().upper()
            for x in sorted_items
            if str(x).strip()
        ]
    else:
        st.warning("streamlit-sortables not installed. Install it locally and in requirements.txt.")
        order_text = st.text_area(
            "Enter owner order (comma or newline separated)",
            value="\n".join(st.session_state.draft_order),
            height=170,
            key="draft_order_textarea",
        )
        new_order = [x.strip().upper() for x in re.split(r"[,\n]+", order_text) if x.strip()]

    if st.button("Submit (reset draft)", key="submit_draft_order"):
        if len(new_order) != settings.num_teams:
            st.error(f"Need exactly {settings.num_teams} owners. You entered {len(new_order)}.")
        else:
            reset_draft(pool, new_order)
            st.rerun()


# Title
st.title("DraftBot 9000")

# Compute VORP and availability
pool_vorp, repl = compute_vorp_with_flex(pool, settings)
pool_vorp["available"] = pool_vorp["player_id"].map(st.session_state.availability).fillna(True)

# Header line controls (all on one line)
pick = next_pick_number(st.session_state.draft_log)
rnd = pick_to_round(pick, settings.num_teams)
my_next_pick = next_pick_for_owner(
    start_pick=pick,
    draft_order=st.session_state.draft_order,
    owner=st.session_state.my_owner,
    num_teams=settings.num_teams,
    max_picks=MAX_PICKS,
)
picks_until_my_turn = None if my_next_pick is None else int(my_next_pick - pick)

draft_complete = pick > MAX_PICKS

display_round = min(rnd, 14)
display_pick = min(pick, MAX_PICKS)

h1, h2, h3, h4, h5, h6 = st.columns([0.9, 1.1, 1.6, 3.0, 1.2, 1.5], vertical_alignment="center")

with h1:
    st.markdown(f"**Round:** {display_round}")
with h2:
    st.markdown(f"**Overall Pick:** {display_pick}")
with h3:
    if picks_until_my_turn is None:
        st.markdown("**Picks Until Your Turn:** -")
    else:
        st.markdown(f"**Picks Until Your Turn:** {picks_until_my_turn}")
with h4:
    st.session_state.current_owner = st.selectbox(
        "Currently picking owner",
        st.session_state.draft_order,
        index=st.session_state.draft_order.index(st.session_state.current_owner),
        label_visibility="collapsed",
        disabled=draft_complete,
    )
with h5:
    st.button("Undo last pick", on_click=undo_last_pick, disabled=st.session_state.draft_log.empty)
with h6:
    if st.button("Advance to next owner", disabled=draft_complete):
        next_p = next_pick_number(st.session_state.draft_log)
        if next_p <= MAX_PICKS:
            idx = snake_team_index(next_p, settings.num_teams)
            st.session_state.current_owner = st.session_state.draft_order[idx]


if draft_complete:
    st.success("Congratulations on finishing this year's draft")

# Main layout
left, right = st.columns([1.55, 1.0], gap="large")

# Main layout
left, right = st.columns([1.25, 1.0], gap="large")

# LEFT: quick pick + filters (keep this part light)
with left:
    st.markdown("#### Quick pick (type to search)")
    # keep your existing quick pick block here (qp1/qp2 etc)

    st.divider()

    st.subheader("Player Bubbles")
    search = st.text_input("Search player (filters lists below)", value="", disabled=draft_complete)
    sort_by = st.selectbox("Sort by", ["VORP", "Projected Points"], index=0, disabled=draft_complete)
    show_n = st.slider("Show top N per position", 10, 90, 35)

# RIGHT: put ALL right-panel stuff here so it appears next to Player Bubbles
with right:
    dl = st.session_state.draft_log.copy()

    weighted_histories = []
    for sheet_name, w in DRAFT_RESULTS_SHEETS_WEIGHTS:
        if sheet_name in history_blocks:
            weighted_histories.append((history_blocks[sheet_name], w))

    sev = run_risk(settings, weighted_histories, dl, pool_vorp)

    # Put MC Best Pick + Best Pick Now + Best Available + Run Risk + Draft Signals Map + My Roster here
    # IMPORTANT: make sure your entire "if on_clock:" MC block is indented inside this right column

    # Example ordering
    # 1) MC Best Pick / Best pick now
    # 2) Best available by position
    # 3) Run risk (with 1 to 2 sentence blurb shown once)
    # 4) Draft signals map
    # 5) My roster
    # 6) Draft Log + Draft Analysis (optional, can stay below)

# LEFT again: now render the heavy tabs and buttons
with left:
    tabs = st.tabs(POSITIONS)
    for i, pos in enumerate(POSITIONS):
        with tabs[i]:
            sub = pool_vorp[(pool_vorp["pos"] == pos) & (pool_vorp["available"])].copy()
            if search.strip():
                sub = sub[sub["player"].str.contains(search.strip(), case=False, na=False)]
            if sort_by == "VORP":
                sub = sub.sort_values("vorp", ascending=False)
            else:
                sub = sub.sort_values("proj_points", ascending=False)

            sub = sub.head(show_n)

            cols = st.columns(3)
            for j, (_, rrow) in enumerate(sub.iterrows()):
                label = f"{rrow['player']} ({rrow['team']}) | {pos} | VORP {rrow['vorp']:.1f}"
                with cols[j % 3]:
                    if st.button(label, key=f"{pos}_{rrow['player_id']}", disabled=draft_complete):
                        draft_player(rrow, st.session_state.current_owner)
                        st.rerun()


    # Best pick now above best available
    on_clock = (st.session_state.current_owner == st.session_state.my_owner) and (not draft_complete)

    if on_clock:
        key = mc_on_clock_key(
            draft_log=st.session_state.draft_log,
            current_pick=pick,
            my_owner=st.session_state.my_owner,
            draft_order=st.session_state.draft_order,
        )

    rerun_clicked = st.button("Re-run Monte Carlo now")

    if rerun_clicked or (st.session_state.get("mc_last_key") != key):
        with st.spinner("Running Monte Carlo simulations for your pick..."):
            mc_df = mc_best_pick_on_clock(
                pool_vorp=pool_vorp,
                draft_order=st.session_state.draft_order,
                draft_log=st.session_state.draft_log,
                settings=settings,
                current_pick=pick,
                my_owner=st.session_state.my_owner,
                sims=int(st.session_state.get("mc_sims", MC_SIMS_DEFAULT)),
                candidates=int(st.session_state.get("mc_candidates", MC_CANDIDATES_DEFAULT)),
            )
        st.session_state.mc_last_key = key
        st.session_state.mc_last_df = mc_df
    else:
        mc_df = st.session_state.get("mc_last_df", pd.DataFrame())

    if mc_df is not None and not mc_df.empty:
        top = mc_df.iloc[0]
        st.success(
            f"MC Best Pick: {top['player']} ({top['pos']}) | EV {top['ev_score']:.1f} | VORP {top['vorp']:.1f}"
        )
        st.dataframe(mc_df, use_container_width=True)
    else:
        st.info("Monte Carlo returned no results. Falling back to greedy ranking.")
        # fall through to your existing greedy logic below

    st.markdown("### Best pick now")

    on_clock = (st.session_state.current_owner == st.session_state.my_owner) and (not draft_complete)

    # Always define sims/cands + key BEFORE using key
    sims = int(st.session_state.get("mc_sims", MC_SIMS_DEFAULT))
    cands = int(st.session_state.get("mc_candidates", MC_CANDIDATES_DEFAULT))

    key = mc_on_clock_key(
        draft_log=st.session_state.draft_log,
        current_pick=pick,
        my_owner=st.session_state.my_owner,
        draft_order=st.session_state.draft_order,
        sims=sims,
        candidates=cands,
    )

    rerun_clicked = st.button("Re-run Monte Carlo now", disabled=not on_clock)

    if on_clock:
        # key is guaranteed to exist here
        if rerun_clicked or (st.session_state.get("mc_last_key") != key):
            with st.spinner("Running Monte Carlo simulations for your pick..."):
                mc_df = mc_best_pick_on_clock(
                    pool_vorp=pool_vorp,
                    draft_order=st.session_state.draft_order,
                    draft_log=st.session_state.draft_log,
                    settings=settings,
                    current_pick=pick,
                    my_owner=st.session_state.my_owner,
                    sims=sims,
                    candidates=cands,
                )
            st.session_state.mc_last_key = key
            st.session_state.mc_last_df = mc_df
        else:
            mc_df = st.session_state.get("mc_last_df", pd.DataFrame())

        if mc_df is not None and not mc_df.empty:
            top = mc_df.iloc[0]
            st.success(f"MC Best Pick: {top['player']} ({top['pos']}) | EV {top['ev_score']:.1f}")
            st.dataframe(mc_df, use_container_width=True)
        else:
            st.info("No Monte Carlo results (yet).")
    else:
        st.info("Monte Carlo runs when you are on the clock.")



    st.markdown("### Best available by position")
    panel = best_available_by_pos(pool_vorp)
    st_df(panel, hide_index=True)

    st.markdown("### My roster")
    my_dl = dl[dl["owner"] == st.session_state.my_owner].sort_values("pick")
    if my_dl.empty:
        st.caption("No picks yet.")
    else:
        st_df(my_dl[["pick", "round", "player", "pos", "team", "proj_points", "vorp"]], hide_index=True)

    # Run risk: position and bar on same line
    st.markdown("### Run risk")
    st.caption(
        "Run detection estimates whether a position is starting to be drafted in a cluster. "
        "It uses recent picks plus historical round tendencies to warn you when a position may thin out soon."
    )

    for pos in POSITIONS:
        c1, c2 = st.columns([0.22, 0.78], vertical_alignment="center")
        with c1:
            st.write(pos)
        with c2:
            st.progress(int(sev[pos] * 100))

    # Draft signals map below My roster
    st.markdown("### Draft signals map")
    other = dl[dl["owner"] != st.session_state.my_owner].copy()
    if not other.empty:
        pos_map = {p: k for k, p in enumerate(POSITIONS)}
        other["pos_y"] = other["pos"].map(pos_map).astype(float)

        chart = (
            alt.Chart(other)
            .mark_circle(size=80)
            .encode(
                x=alt.X("pick:Q", title="Pick"),
                y=alt.Y(
                    "pos_y:Q",
                    title="",
                    axis=alt.Axis(
                        values=list(pos_map.values()),
                        labelExpr="['QB','RB','WR','TE','K','DEF'][datum.value]",
                    ),
                ),
                color=alt.Color("pos:N", title="Position"),
                tooltip=["pick", "round", "owner", "player", "pos", "vorp"],
            )
            .properties(height=200)
        )
        st_altair(chart)
    else:
        st.info("No picks logged yet.")

st.divider()

st.subheader("Draft Log")
st_df(st.session_state.draft_log.sort_values("pick"), hide_index=True)

st.caption(
    f"Replacement points: QB {repl['QB']:.1f} | RB {repl['RB']:.1f} | WR {repl['WR']:.1f} | "
    f"TE {repl['TE']:.1f} | FLEX {repl['FLEX_RBWR']:.1f} | K {repl['K']:.1f} | DEF {repl['DEF']:.1f}"
)

st.subheader("Draft Analysis")
analysis = compute_team_analysis(st.session_state.draft_log, st.session_state.draft_order)
if analysis.empty:
    st.caption("Draft Analysis will appear after picks are made.")
else:
    st_df(analysis, hide_index=True)
    st.caption("Ranking is based primarily on starter strength (VORP and projected points), with a small depth adjustment.")

